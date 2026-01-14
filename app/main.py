from fastapi import FastAPI, File, UploadFile
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import JSONResponse
import re
import io
from typing import List, Dict, Any, Optional

app = FastAPI()

# --- CORS (כדי ש-Base44 יוכל לקרוא) ---
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# -----------------------------
# Healthcheck
# -----------------------------
@app.get("/")
def root():
    return {"status": "ok", "service": "LTG Questionnaire Extractor"}

# -----------------------------
# Helpers
# -----------------------------

QUESTION_PATTERNS = [
    r"^\s*\d+[\)\.\-]\s+",      # 1) / 1. / 1-
    r"^\s*[\u2022\-\*]\s+",     # bullet
    r"^\s*[א-ת]\)\s+",          # א) ב) ...
]

ANSWER_PREFIX_PATTERNS = [
    r"^\s*[\u2022\-\*]\s+",     # • / - / *
    r"^\s*\d+[\)\.\-]\s+",      # 1) / 1. / 1-
    r"^\s*[א-ת]\)\s+",          # א) ב) ...
]

MULTI_CHOICE_HINTS = [
    "אפשר לבחור", "יותר מתשובה אחת", "מספר תשובות", "בחר/י עד", "בחרו עד", "סמן/י את כל"
]

def clean_text(s: str) -> str:
    if not s:
        return ""
    s = s.replace("\u00a0", " ")
    s = re.sub(r"\s+", " ", s).strip()
    return s

def looks_like_question(line: str) -> bool:
    t = clean_text(line)
    if len(t) < 3:
        return False
    # סימן שאלה או נקודתיים בסוף (הרבה שאלונים עם ":")
    if t.endswith("?") or re.search(r":\s*$", t):
        return True
    # מתחיל במספור
    for p in QUESTION_PATTERNS:
        if re.search(p, t):
            # לא כל מספור הוא שאלה, אבל אם יש גם "?" או ":" או ארוך מספיק – נחשב
            if "?" in t or re.search(r":\s*$", t) or len(t) >= 8:
                return True
    # ניסוח שאלה בעברית נפוץ (היוריסטיקה עדינה)
    if re.search(r"\b(עד כמה|באיזו מידה|מה דעתך|למי תצביע|האם|כמה)\b", t):
        return True
    return False

def strip_prefixes(s: str) -> str:
    t = s
    # מסיר מספור/בולטים מההתחלה
    for p in QUESTION_PATTERNS + ANSWER_PREFIX_PATTERNS:
        t = re.sub(p, "", t)
    return clean_text(t)

def looks_like_answer(line: str) -> bool:
    t = clean_text(line)
    if len(t) < 1:
        return False
    # תשובה בדרך כלל מגיעה כבולט/מספור
    for p in ANSWER_PREFIX_PATTERNS:
        if re.search(p, t):
            return True
    # או מילת תשובה קצרה מאוד אחרי שאלה
    if len(t) <= 40 and not looks_like_question(t) and not t.endswith("?"):
        return True
    return False

def infer_question_type(question_text: str, answers: List[str]) -> str:
    qt = clean_text(question_text)
    if not answers:
        return "open"
    for hint in MULTI_CHOICE_HINTS:
        if hint in qt:
            return "multi_choice"
    # ברירת מחדל: סגורה יחידה
    return "single_choice"

# -----------------------------
# DOCX parsing
# -----------------------------
def parse_docx_questions(file_bytes: bytes) -> List[Dict[str, Any]]:
    try:
        from docx import Document  # python-docx
    except Exception as e:
        raise RuntimeError("python-docx is not installed. Add it to requirements.txt") from e

    doc = Document(io.BytesIO(file_bytes))
    paragraphs = [clean_text(p.text) for p in doc.paragraphs]
    paragraphs = [p for p in paragraphs if p]  # remove empty

    questions: List[Dict[str, Any]] = []
    current_q: Optional[Dict[str, Any]] = None

    def flush_current():
        nonlocal current_q
        if not current_q:
            return
        # ניקוי תשובות + הסרת כפילויות רציפות
        answers_clean = []
        seen = set()
        for a in current_q.get("answers", []):
            aa = strip_prefixes(a)
            if not aa:
                continue
            if aa in seen:
                continue
            seen.add(aa)
            answers_clean.append(aa)

        qtext = strip_prefixes(current_q.get("text", ""))
        qtype = infer_question_type(qtext, answers_clean)

        questions.append({
            "text": qtext,
            "type": qtype,
            "answers": answers_clean,
            "meta": {
                "source": "docx",
                "question_index": len(questions) + 1
            }
        })
        current_q = None

    i = 0
    while i < len(paragraphs):
        line = paragraphs[i]

        # אם זה נראה כמו שאלה -> פותחים שאלה חדשה
        if looks_like_question(line):
            flush_current()
            current_q = {"text": line, "answers": []}
            i += 1
            # אוספים תשובות "עד השאלה הבאה"
            while i < len(paragraphs):
                nxt = paragraphs[i]
                if looks_like_question(nxt):
                    break
                # אם זה נראה כמו תשובה – נוסיף
                if looks_like_answer(nxt):
                    current_q["answers"].append(nxt)
                else:
                    # טקסט המשך של השאלה (לעיתים שאלה נשברת לשתי שורות)
                    # נוסיף אותו לשאלה אם אין עדיין תשובות
                    if current_q and not current_q["answers"]:
                        current_q["text"] = clean_text(current_q["text"] + " " + nxt)
                i += 1
            continue

        # אם אין שאלה פתוחה ועדיין יש טקסט — נדלג
        i += 1

    flush_current()
    return questions

# -----------------------------
# CSV / XLSX basic parsing (optional)
# -----------------------------
def parse_csv_questions(file_bytes: bytes) -> List[Dict[str, Any]]:
    import csv
    text = file_bytes.decode("utf-8-sig", errors="replace")
    reader = csv.DictReader(io.StringIO(text))
    out = []
    for row in reader:
        q = clean_text(row.get("question") or row.get("Question") or row.get("שאלה") or "")
        if not q:
            continue
        answers_raw = row.get("answers") or row.get("Answers") or row.get("תשובות") or ""
        answers = [clean_text(x) for x in re.split(r"[;\|]", answers_raw) if clean_text(x)]
        out.append({
            "text": q,
            "type": infer_question_type(q, answers),
            "answers": answers,
            "meta": {"source": "csv", "question_index": len(out) + 1}
        })
    return out

def parse_xlsx_questions(file_bytes: bytes) -> List[Dict[str, Any]]:
    try:
        import openpyxl
    except Exception as e:
        raise RuntimeError("openpyxl is not installed. Add it to requirements.txt") from e

    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    ws = wb.active

    # מניחים שורה 1 = כותרות
    headers = [clean_text(str(c.value)) if c.value is not None else "" for c in ws[1]]
    idx_q = None
    idx_a = None
    for j, h in enumerate(headers):
        if h.lower() in ["question", "שאלה"]:
            idx_q = j
        if h.lower() in ["answers", "תשובות"]:
            idx_a = j

    if idx_q is None:
        return []

    out = []
    for r in ws.iter_rows(min_row=2, values_only=True):
        q = clean_text(str(r[idx_q])) if r[idx_q] is not None else ""
        if not q:
            continue
        answers = []
        if idx_a is not None and r[idx_a] is not None:
            answers = [clean_text(x) for x in re.split(r"[;\|]", str(r[idx_a])) if clean_text(x)]
        out.append({
            "text": q,
            "type": infer_question_type(q, answers),
            "answers": answers,
            "meta": {"source": "xlsx", "question_index": len(out) + 1}
        })
    return out

# -----------------------------
# API endpoint
# -----------------------------
@app.post("/extract")
async def extract_questions(file: UploadFile = File(...)):
    raw = await file.read()
    filename = file.filename or "uploaded"
    content_type = file.content_type or ""

    ext = ""
    if "." in filename:
        ext = filename.rsplit(".", 1)[-1].lower()

    try:
        if ext == "docx" or content_type == "application/vnd.openxmlformats-officedocument.wordprocessingml.document":
            questions = parse_docx_questions(raw)
        elif ext == "csv" or "csv" in content_type:
            questions = parse_csv_questions(raw)
        elif ext in ["xlsx", "xlsm", "xltx", "xltm"] or "spreadsheet" in content_type:
            questions = parse_xlsx_questions(raw)
        else:
            return JSONResponse(
                status_code=400,
                content={
                    "status": "error",
                    "message": f"Unsupported file type. Please upload DOCX/CSV/XLSX.",
                    "filename": filename,
                    "content_type": content_type,
                },
            )

        return JSONResponse(
            {
                "status": "parsed",
                "filename": filename,
                "content_type": content_type,
                "questions_count": len(questions),
                "questions": questions,
            }
        )

    except Exception as e:
        return JSONResponse(
            status_code=500,
            content={
                "status": "error",
                "filename": filename,
                "content_type": content_type,
                "message": str(e),
            },
        )
